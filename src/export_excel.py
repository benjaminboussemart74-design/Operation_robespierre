"""
Génère un fichier Excel formaté à partir des résultats éolien Somme.
Un onglet Synthèse + un onglet par député, avec codes couleurs et liens AN.

Usage : python3 src/export_excel.py
Prérequis : pip install openpyxl
"""

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink
except ImportError:
    print("❌ openpyxl manquant. Lance : pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).parent.parent
RESULTS_DIR = BASE_DIR / "results"

# ── Couleurs ──────────────────────────────────────────────────────────────────
C_FAVORABLE  = "C6EFCE"   # vert clair
C_NUANCE     = "FFEB9C"   # jaune
C_CRITIQUE   = "FFC7CE"   # rose/rouge
C_HEADER_BG  = "1F3864"   # bleu marine
C_HEADER_FG  = "FFFFFF"
C_ALT_ROW    = "EEF2F7"   # gris très clair pour lignes paires
C_SOURCE = {
    "Amendement":      "DDEEFF",
    "Séance plénière": "FFEEDD",
    "Question écrite": "EEFFDD",
}
C_GROUPE = {
    "RN":      "003189",
    "LFI-NFP": "BB2020",
    "ECOS":    "2E7D32",
    "RE":      "FFD700",
}

THEME_LABELS = {
    "eolien_infrastructure": "Infrastructure",
    "eolien_administratif":  "Réglementation",
    "eolien_impacts":        "Impacts",
    "eolien_politique":      "Politique",
}

# ── URLs Assemblée nationale ──────────────────────────────────────────────────
def url_for(row: dict) -> str:
    uid = row.get("uid", "") or row.get("Référence", "")
    src = row.get("source", "") or row.get("Source", "")
    if src == "Question écrite" and uid.startswith("QANR"):
        return f"https://www.assemblee-nationale.fr/dyn/17/questions/detail/QE/{uid}"
    if src == "Amendement":
        texte_ref = row.get("texte_ref", "")
        if texte_ref and uid:
            return f"https://www.assemblee-nationale.fr/dyn/17/amendements/{texte_ref}/{uid}"
    if src == "Séance plénière" and uid.startswith("CRS"):
        # uid format: CRSANR5L17S2025O1N014 → l17S2025O1N014
        slug = uid.replace("CRSANR5", "").lower() if uid.startswith("CRSANR5") else uid.lower()
        return f"https://www.assemblee-nationale.fr/dyn/17/comptes-rendus/{slug}"
    return ""


# ── Helpers styles ────────────────────────────────────────────────────────────
def header_fill():
    return PatternFill("solid", fgColor=C_HEADER_BG)

def header_font():
    return Font(bold=True, color=C_HEADER_FG, size=11)

def cell_font(bold=False, size=10):
    return Font(bold=bold, size=size)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def position_fill(pos: str):
    if "Favorable" in pos:
        return PatternFill("solid", fgColor=C_FAVORABLE)
    if "Critique" in pos or "Opposé" in pos:
        return PatternFill("solid", fgColor=C_CRITIQUE)
    return PatternFill("solid", fgColor=C_NUANCE)

def source_fill(src: str):
    return PatternFill("solid", fgColor=C_SOURCE.get(src, "FFFFFF"))

def set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

def apply_header(ws, headers, widths):
    ws.append(headers)
    row = ws.max_row
    for i, w in enumerate(widths, 1):
        c = ws.cell(row, i)
        c.fill = header_fill()
        c.font = header_font()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
        set_col_width(ws, i, w)
    ws.row_dimensions[row].height = 30


# ── Onglet Synthèse ───────────────────────────────────────────────────────────
def build_synthese(wb, synthese: list[dict]):
    ws = wb.active
    ws.title = "Synthèse"
    ws.sheet_view.showGridLines = False

    # Titre
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "ÉOLIEN — DÉPUTÉS DE LA SOMME — 17e LÉGISLATURE"
    title_cell.font = Font(bold=True, size=14, color=C_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.append([])  # ligne vide

    headers = ["Député·e", "Groupe", "Circo", "Amendements", "Séances", "QE", "Total", "Thème dominant", "Position"]
    widths  = [22,         10,       6,       13,            9,         6,    8,       22,               22]
    apply_header(ws, headers, widths)

    for i, row in enumerate(sorted(synthese, key=lambda x: -x["Total"])):
        pos = row.get("Position déduite", "")
        values = [
            row["Député·e"], row["Groupe"], row["Circo Somme"],
            row["Amendements"], row["Séances"], row["QE"], row["Total"],
            row["Thème dominant"], pos,
        ]
        ws.append(values)
        r = ws.max_row
        fill = position_fill(pos)
        for col in range(1, 10):
            c = ws.cell(r, col)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.font = cell_font(size=10)
            if col == 9:
                c.fill = fill
                c.font = Font(bold=True, size=10)
            elif i % 2 == 1:
                c.fill = PatternFill("solid", fgColor=C_ALT_ROW)

        # Couleur groupe
        gc = ws.cell(r, 2)
        grp = row["Groupe"]
        if grp in C_GROUPE:
            gc.font = Font(bold=True, color=C_GROUPE[grp], size=10)

        ws.row_dimensions[r].height = 20

    # Légende
    ws.append([])
    ws.append(["Légende"])
    ws["A" + str(ws.max_row)].font = Font(bold=True, color=C_HEADER_BG)
    for label, color in [("✓ Favorable", C_FAVORABLE), ("~ Nuancé / Mixte", C_NUANCE), ("⚠ Critique / Opposé", C_CRITIQUE)]:
        ws.append([label])
        c = ws.cell(ws.max_row, 1)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(bold=True, size=9)

    ws.freeze_panes = "A4"


# ── Onglet par député ─────────────────────────────────────────────────────────
def build_depute_sheet(wb, nom: str, groupe: str, rows: list[dict]):
    safe_name = nom[:28].replace("/", "-").replace(":", "-")
    ws = wb.create_sheet(title=safe_name)
    ws.sheet_view.showGridLines = False

    # Titre onglet
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = f"{nom} — {groupe}"
    color = C_GROUPE.get(groupe, C_HEADER_BG)
    t.fill = PatternFill("solid", fgColor=color)
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.append([])

    headers = ["Source", "Date", "Rôle", "Thèmes", "Extrait", "Référence", "Lien AN"]
    widths  = [14,       12,     12,     28,        60,        20,          40]
    apply_header(ws, headers, widths)

    for i, row in enumerate(sorted(rows, key=lambda x: x.get("date", ""))):
        src   = row.get("source", "")
        date  = row.get("date", "")
        role  = row.get("role", "")
        themes = " | ".join(THEME_LABELS.get(t, t) for t in row.get("themes", []))
        extrait = row.get("extrait", "")[:300]
        uid   = row.get("uid", "")
        lien  = url_for(row)

        ws.append([src, date, role, themes, extrait, uid, lien if lien else ""])
        r = ws.max_row
        sfill = source_fill(src)

        for col in range(1, 8):
            c = ws.cell(r, col)
            c.border = thin_border()
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = cell_font(size=9)
            if col == 1:
                c.fill = sfill
                c.font = Font(bold=True, size=9)
            elif i % 2 == 1 and col != 7:
                c.fill = PatternFill("solid", fgColor=C_ALT_ROW)

        # Hyperlink sur la cellule Lien AN
        if lien:
            link_cell = ws.cell(r, 7)
            link_cell.hyperlink = lien
            link_cell.value = "→ Voir sur AN"
            link_cell.font = Font(color="0563C1", underline="single", size=9)

        ws.row_dimensions[r].height = 60

    ws.freeze_panes = "A4"


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    json_path = RESULTS_DIR / "eolien_somme_latest.json"
    if not json_path.exists():
        print(f"❌ Fichier introuvable : {json_path}")
        print("   Lance d'abord : python3 src/extract_eolien.py")
        sys.exit(1)

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    synthese = data["synthese"]
    detail   = data["detail"]

    # Reconstruire le détail brut depuis extract (pour avoir uid, texte_ref, source brut)
    # On relit directement le JSON complet produit par extract_eolien
    # Le detail dans eolien_somme_latest.json a les champs renommés en FR
    # → on recharge les données brutes
    sys.path.insert(0, str(Path(__file__).parent))
    from extract_eolien import extract_amendments_eolien, extract_cr_eolien, extract_qe_eolien, DEPUTES_SOMME

    print("Rechargement des données brutes...")
    all_results = extract_amendments_eolien() + extract_cr_eolien() + extract_qe_eolien()
    print(f"  → {len(all_results)} occurrences")

    wb = openpyxl.Workbook()
    build_synthese(wb, synthese)

    for ref, dep in DEPUTES_SOMME.items():
        nom = f"{dep['prenom']} {dep['nom']}"
        groupe = dep["groupe"]
        rows = [r for r in all_results if r["acteur_ref"] == ref]
        build_depute_sheet(wb, nom, groupe, rows)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = RESULTS_DIR / f"eolien_somme_{ts}.xlsx"
    out_latest = RESULTS_DIR / "eolien_somme_latest.xlsx"
    wb.save(out)
    wb.save(out_latest)
    print(f"✓ Excel créé : {out_latest}")
    print(f"  Ouverture...")

    import subprocess, platform
    if platform.system() == "Darwin":
        subprocess.run(["open", str(out_latest)])


if __name__ == "__main__":
    main()
